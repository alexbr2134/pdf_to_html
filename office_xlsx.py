"""
Excel (.xlsx) → блоки: prose + отдельные таблицы на листе.

Смотрим значения, merges и borders. .xls сюда не заходит — его роутят раньше.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from office_table_regions import (
    find_trailing_chrome_start,
    grid_rows_as_prose_lines,
    should_emit_as_html_table,
    trim_trailing_form_chrome,
)
from pdf_html_pipeline import Cell


BlockKind = Literal["paragraph", "table"]


@dataclass
class ExcelBlock:
    kind: BlockKind
    text: str = ""
    grid: list[list[Cell]] | None = None


@dataclass
class ExcelSheetModel:
    name: str
    blocks: list[ExcelBlock] = field(default_factory=list)
    text: str = ""
    n_rows: int = 0
    n_cols: int = 0
    n_nonempty: int = 0
    n_tables: int = 0
    hidden: bool = False
    grid: list[list[Cell]] = field(default_factory=list)  # первая таблица, для старого кода


@dataclass
class ExcelWorkbookModel:
    sheets: list[ExcelSheetModel] = field(default_factory=list)
    text: str = ""
    warnings: list[str] = field(default_factory=list)
    source_format: str = "xlsx"


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else "нет"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.10f}".rstrip("0").rstrip(".")
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _cell_has_border(cell) -> bool:
    b = getattr(cell, "border", None)
    if b is None:
        return False
    for side in ("left", "right", "top", "bottom"):
        edge = getattr(b, side, None)
        if edge is not None and getattr(edge, "style", None):
            return True
    return False


def _build_grid_from_matrix(
    matrix: list[list[str]],
    merges: list[tuple[int, int, int, int]],
) -> list[list[Cell]]:
    """matrix + merges (r0,c0,r1,c1) relative → Cell grid with spans."""
    if not matrix:
        return []
    n_rows = len(matrix)
    n_cols = max((len(r) for r in matrix), default=0)
    for row in matrix:
        if len(row) < n_cols:
            row.extend([""] * (n_cols - len(row)))

    covered = [[False] * n_cols for _ in range(n_rows)]
    span_map: dict[tuple[int, int], tuple[int, int]] = {}
    for r0, c0, r1, c1 in merges:
        if r0 < 0 or c0 < 0 or r0 >= n_rows or c0 >= n_cols:
            continue
        rs = max(1, min(n_rows - 1, r1) - r0 + 1)
        cs = max(1, min(n_cols - 1, c1) - c0 + 1)
        span_map[(r0, c0)] = (rs, cs)
        for r in range(r0, r0 + rs):
            for c in range(c0, c0 + cs):
                if (r, c) != (r0, c0) and r < n_rows and c < n_cols:
                    covered[r][c] = True

    grid: list[list[Cell]] = []
    for r in range(n_rows):
        row_cells: list[Cell] = []
        for c in range(n_cols):
            if covered[r][c]:
                row_cells.append(
                    Cell(
                        row=r,
                        col=c,
                        bbox=None,
                        text="",
                        covered=True,
                        is_placeholder=True,
                    )
                )
                continue
            rs, cs = span_map.get((r, c), (1, 1))
            text = matrix[r][c]
            row_cells.append(
                Cell(
                    row=r,
                    col=c,
                    bbox=None,
                    text=text,
                    rowspan=rs,
                    colspan=cs,
                    covered=False,
                    is_placeholder=not bool(text),
                )
            )
        grid.append(row_cells)
    return grid


def _compact_grid(grid: list[list[Cell]]) -> list[list[Cell]]:
    """Убирает полностью пустые строки/колонки (без учёта covered-слотов)."""
    if not grid:
        return grid
    n_rows = len(grid)
    n_cols = len(grid[0])

    keep_cols = []
    for c in range(n_cols):
        useful = False
        for r in range(n_rows):
            cell = grid[r][c]
            if cell.covered:
                continue
            if (cell.text or "").strip() or cell.colspan > 1 or cell.rowspan > 1:
                useful = True
                break
        if useful:
            keep_cols.append(c)
    if not keep_cols:
        return []

    # remap: old col → new col
    col_map = {old: new for new, old in enumerate(keep_cols)}
    keep_col_set = set(keep_cols)

    new_grid: list[list[Cell]] = []
    for r, row in enumerate(grid):
        # drop fully empty rows
        row_has = False
        for c in keep_cols:
            cell = row[c]
            if cell.covered:
                continue
            if (cell.text or "").strip() or cell.colspan > 1:
                row_has = True
                break
        if not row_has:
            continue

        new_row: list[Cell] = []
        for c in keep_cols:
            cell = row[c]
            if cell.covered:
                # keep covered only if master column kept
                new_row.append(
                    Cell(
                        row=len(new_grid),
                        col=col_map[c],
                        bbox=None,
                        text="",
                        covered=True,
                        is_placeholder=True,
                    )
                )
                continue
            # shrink colspan if it spans dropped columns
            cs = cell.colspan
            if cs > 1:
                spanned = [c + k for k in range(cs) if c + k < n_cols]
                kept = [x for x in spanned if x in keep_col_set]
                cs = max(1, len(kept))
            new_row.append(
                Cell(
                    row=len(new_grid),
                    col=col_map[c],
                    bbox=None,
                    text=cell.text,
                    rowspan=cell.rowspan,
                    colspan=cs,
                    covered=False,
                    is_placeholder=not bool((cell.text or "").strip()),
                )
            )
        new_grid.append(new_row)

    # fix rowspan after row drops: clamp
    for r, row in enumerate(new_grid):
        for cell in row:
            if cell.covered:
                continue
            if cell.rowspan > 1:
                cell.rowspan = min(cell.rowspan, len(new_grid) - r)
    return new_grid


def _row_is_index_header(values: list[str]) -> bool:
    """Нумерация колонок: 1..N (ТОРГ) или А 1 1а 1б… (СФ/УПД)."""
    nums: list[int] = []
    for v in values:
        t = (v or "").strip()
        if t.isdigit() and 1 <= int(t) <= 40:
            nums.append(int(t))
        elif t:
            if nums:
                break
    if (
        len(nums) >= 6
        and nums[0] == 1
        and nums == list(range(1, len(nums) + 1))
    ):
        return True

    parts = [
        (v or "").strip().lower().replace("\n", "").replace(" ", "")
        for v in values
        if (v or "").strip()
    ]
    if len(parts) < 6:
        return False
    short = [p for p in parts if len(p) <= 3]
    if len(short) < 6:
        return False
    has_letter_codes = any(re.fullmatch(r"\d{1,2}[аaбb]", p) for p in parts)
    starts_with_a = parts[0] in {"а", "a"}
    return "1" in parts and "2" in parts and (has_letter_codes or starts_with_a)


def _detect_regions(
    n_rows: int,
    row_has_value: list[bool],
    row_border_count: list[int],
    row_merge_count: list[int],
    row_values: list[list[str]],
) -> list[tuple[int, int]]:
    """
    Возвращает список (r0, r1) inclusive 0-based регионов-таблиц.

    Эвристики:
    - строки с бордерами / merges / плотным контентом — «табличные»;
    - пустые / почти пустые без бордеров — разделители;
    - INDEX-строка (1..N) закрепляет data-таблицу.
    """
    if n_rows <= 0:
        return []

    score = [0.0] * n_rows
    for r in range(n_rows):
        if row_has_value[r]:
            score[r] += 1.0
        if row_border_count[r] >= 3:
            score[r] += 2.0
        elif row_border_count[r] >= 1:
            score[r] += 0.8
        if row_merge_count[r] >= 1:
            score[r] += 0.8
        if _row_is_index_header(row_values[r]):
            score[r] += 3.0

    is_tab = [
        score[r] >= 1.5 or (row_border_count[r] >= 2 and row_has_value[r])
        for r in range(n_rows)
    ]

    # подтянуть соседние строки с текстом к bordered-блоку
    for r in range(n_rows):
        if not is_tab[r] and row_has_value[r]:
            neigh = (r > 0 and is_tab[r - 1]) or (r + 1 < n_rows and is_tab[r + 1])
            if neigh and (row_border_count[r] >= 1 or row_merge_count[r] >= 1):
                is_tab[r] = True

    regions: list[tuple[int, int]] = []
    r = 0
    while r < n_rows:
        if not is_tab[r]:
            r += 1
            continue
        start = r
        while r < n_rows and is_tab[r]:
            r += 1
        end = r - 1
        span = end - start + 1
        border_sum = sum(row_border_count[i] for i in range(start, end + 1))
        value_rows = sum(1 for i in range(start, end + 1) if row_has_value[i])
        # одиночный заголовок без рамок — не таблица
        if span == 1 and border_sum < 2 and value_rows <= 1:
            continue
        if value_rows == 0 and border_sum < 4:
            continue
        regions.append((start, end))

    # пустой зазор между двумя рамками — это разные блоки, не склеиваем
    merged: list[tuple[int, int]] = []
    for reg in regions:
        if not merged:
            merged.append(reg)
            continue
        prev_s, prev_e = merged[-1]
        s, e = reg
        gap = s - prev_e - 1
        gap_border = (
            sum(row_border_count[i] for i in range(prev_e + 1, s)) if gap > 0 else 0
        )
        gap_values = (
            sum(1 for i in range(prev_e + 1, s) if row_has_value[i]) if gap > 0 else 0
        )
        if (
            gap == 0
            or (
                gap == 1
                and gap_border >= 2
                and gap_values >= 1
                and sum(row_border_count[i] for i in range(prev_s, prev_e + 1)) >= 4
                and sum(row_border_count[i] for i in range(s, e + 1)) >= 4
            )
        ):
            merged[-1] = (prev_s, e)
        else:
            merged.append(reg)

    final: list[tuple[int, int]] = []
    for s, e in merged:
        index_rows = [
            r for r in range(s, e + 1) if _row_is_index_header(row_values[r])
        ]
        pieces: list[tuple[int, int]] = []
        if len(index_rows) == 1 and index_rows[0] - s >= 3:
            # шапка формы сверху, товарная сетка с index-строки
            idx = index_rows[0]
            data_start = max(s, idx - 3)
            if s <= data_start - 1:
                pieces.append((s, data_start - 1))
            pieces.append((data_start, e))
        else:
            cut = None
            for i in range(s + 2, e - 1):
                if row_border_count[i] == 0 and not row_has_value[i]:
                    if row_border_count[i + 1] == 0 and not row_has_value[i + 1]:
                        above = sum(row_border_count[j] for j in range(s, i))
                        below = sum(row_border_count[j] for j in range(i + 2, e + 1))
                        if above >= 4 and below >= 4:
                            cut = i + 2
                            break
            if cut is not None and cut <= e:
                pieces.append((s, cut - 1))
                pieces.append((cut, e))
            else:
                pieces.append((s, e))

        for ps, pe in pieces:
            tail = None
            for r in range(ps + 3, pe + 1):
                joined = " ".join(
                    (v or "").strip() for v in row_values[r] if (v or "").strip()
                )
                if not joined:
                    continue
                if re.match(r"^(итого|всего\s+по\s+накладной)\b", joined, re.I):
                    continue
                if re.search(
                    r"товарная\s+накладная\s+имеет\s+приложение|"
                    r"порядковых\s+номеров\s+записей|"
                    r"масса\s+груза\s*\(|всего\s+мест|"
                    r"по\s+доверенности|отпуск\s+груза|"
                    r"приложение\s*\(\s*паспорта|всего\s+отпущено",
                    joined,
                    re.I,
                ):
                    tail = r
                    break
            if tail is not None and tail > ps:
                final.append((ps, tail - 1))
                final.append((tail, pe))
            else:
                final.append((ps, pe))
    return final


def _region_to_grid(
    matrix: list[list[str]],
    merges: list[tuple[int, int, int, int]],
    r0: int,
    r1: int,
    active_cols: list[int] | None = None,
) -> list[list[Cell]]:
    """Вырезает регион строк, компактит колонки, накладывает merges."""
    n_cols = max((len(r) for r in matrix), default=0)
    if active_cols is None:
        # columns with any content or merge coverage in region
        col_useful = [False] * n_cols
        for r in range(r0, r1 + 1):
            row = matrix[r] if r < len(matrix) else []
            for c, val in enumerate(row):
                if val:
                    col_useful[c] = True
        for mr0, mc0, mr1, mc1 in merges:
            if mr1 < r0 or mr0 > r1:
                continue
            for c in range(max(0, mc0), min(n_cols, mc1 + 1)):
                col_useful[c] = True
        active_cols = [c for c, ok in enumerate(col_useful) if ok]
    if not active_cols:
        return []

    col_index = {c: i for i, c in enumerate(active_cols)}
    sub: list[list[str]] = []
    for r in range(r0, r1 + 1):
        row = matrix[r] if r < len(matrix) else []
        sub.append([row[c] if c < len(row) else "" for c in active_cols])

    rel_merges: list[tuple[int, int, int, int]] = []
    for mr0, mc0, mr1, mc1 in merges:
        if mr1 < r0 or mr0 > r1:
            continue
        # clip to region
        rr0 = max(mr0, r0) - r0
        rr1 = min(mr1, r1) - r0
        # map columns: keep only active
        kept = [c for c in range(mc0, mc1 + 1) if c in col_index]
        if not kept:
            continue
        cc0 = col_index[kept[0]]
        cc1 = col_index[kept[-1]]
        rel_merges.append((rr0, cc0, rr1, cc1))

    grid = _build_grid_from_matrix(sub, rel_merges)
    return _compact_grid(grid)


def _prose_from_rows(matrix: list[list[str]], r0: int, r1: int) -> list[str]:
    paras: list[str] = []
    for r in range(r0, r1 + 1):
        if r >= len(matrix):
            break
        parts = [c for c in matrix[r] if c]
        if not parts:
            continue
        # join cells on a row with " | " only if multiple short fields
        if len(parts) == 1:
            paras.append(parts[0])
        elif all(len(p) < 80 for p in parts):
            paras.append(" | ".join(parts))
        else:
            paras.extend(parts)
    return paras


def _sheet_to_blocks(
    matrix: list[list[str]],
    merges: list[tuple[int, int, int, int]],
    row_border_count: list[int],
    row_merge_count: list[int],
) -> tuple[list[ExcelBlock], list[str]]:
    n_rows = len(matrix)
    if n_rows == 0:
        return [], []

    row_has_value = [any(bool(c) for c in row) for row in matrix]
    row_values = matrix
    regions = _detect_regions(
        n_rows, row_has_value, row_border_count, row_merge_count, row_values
    )

    notes: list[str] = []
    if len(regions) > 1:
        notes.append(f"regions={len(regions)}")
    blocks: list[ExcelBlock] = []
    cursor = 0
    for r0, r1 in regions:
        if cursor < r0:
            for para in _prose_from_rows(matrix, cursor, r0 - 1):
                blocks.append(ExcelBlock(kind="paragraph", text=para))
        grid = _region_to_grid(matrix, merges, r0, r1)
        if grid and len(grid) >= 1 and len(grid[0]) >= 1:
            chrome_at = find_trailing_chrome_start(grid)
            tail_grid = None
            if chrome_at is not None and chrome_at >= 3:
                tail_grid = grid[chrome_at:]
                grid = grid[:chrome_at]

            nonempty = sum(
                1
                for row in grid
                for c in row
                if not c.covered and (c.text or "").strip()
            )
            if nonempty == 0:
                pass
            elif should_emit_as_html_table(grid):
                grid = trim_trailing_form_chrome(grid)
                text = "\n".join(
                    c.text
                    for row in grid
                    for c in row
                    if not c.covered and (c.text or "").strip()
                )
                blocks.append(ExcelBlock(kind="table", grid=grid, text=text))
            else:
                for line in grid_rows_as_prose_lines(grid):
                    blocks.append(ExcelBlock(kind="paragraph", text=line))
            if tail_grid:
                for line in grid_rows_as_prose_lines(tail_grid):
                    blocks.append(ExcelBlock(kind="paragraph", text=line))
        cursor = r1 + 1

    if cursor < n_rows:
        for para in _prose_from_rows(matrix, cursor, n_rows - 1):
            blocks.append(ExcelBlock(kind="paragraph", text=para))

    if not blocks:
        used_rows = [i for i, h in enumerate(row_has_value) if h]
        if used_rows:
            grid = _region_to_grid(matrix, merges, used_rows[0], used_rows[-1])
            if grid and should_emit_as_html_table(grid):
                text = "\n".join(
                    c.text
                    for row in grid
                    for c in row
                    if not c.covered and (c.text or "").strip()
                )
                blocks = [ExcelBlock(kind="table", grid=grid, text=text)]
            elif grid:
                for line in grid_rows_as_prose_lines(grid):
                    blocks.append(ExcelBlock(kind="paragraph", text=line))
    return blocks, notes


def extract_xlsx(path: str | Path, *, data_only: bool = True) -> ExcelWorkbookModel:
    """Читает .xlsx/.xlsm через openpyxl, режет лист на таблицы."""
    import openpyxl

    src = Path(path)
    warnings: list[str] = []
    try:
        wb = openpyxl.load_workbook(str(src), data_only=data_only, read_only=False)
    except Exception as exc:
        msg = str(exc).lower()
        if "password" in msg or "encrypt" in msg or "workbook is encrypted" in msg:
            raise PermissionError("encrypted") from exc
        raise

    wb_formulas = None
    if data_only:
        try:
            wb_formulas = openpyxl.load_workbook(str(src), data_only=False)
        except Exception:
            wb_formulas = None

    sheets: list[ExcelSheetModel] = []
    all_text: list[str] = []

    for ws in wb.worksheets:
        hidden = bool(getattr(ws, "sheet_state", "visible") == "hidden")

        # 1) find true used bounds by values (ignore inflated max_row from formatting)
        max_scan_row = min(ws.max_row or 0, 5000)
        max_scan_col = min(ws.max_column or 0, 120)
        last_val_row = 0
        last_val_col = 0
        empty_streak = 0
        for r in range(1, max_scan_row + 1):
            row_hit = False
            for c in range(1, max_scan_col + 1):
                val = ws.cell(r, c).value
                if val not in (None, ""):
                    row_hit = True
                    last_val_row = r
                    last_val_col = max(last_val_col, c)
            if row_hit:
                empty_streak = 0
            else:
                empty_streak += 1
                # stop after long empty tail once we saw content
                if last_val_row > 0 and empty_streak >= 40:
                    break

        if last_val_row == 0:
            sheets.append(
                ExcelSheetModel(name=ws.title, hidden=hidden)
            )
            continue

        # also extend bounds by merges
        merges_abs: list[tuple[int, int, int, int]] = []
        for mr in ws.merged_cells.ranges:
            r0, r1 = mr.min_row, mr.max_row
            c0, c1 = mr.min_col, mr.max_col
            merges_abs.append((r0, c0, r1, c1))
            if r1 <= last_val_row + 5:
                last_val_row = max(last_val_row, min(r1, last_val_row + 30))
                last_val_col = max(last_val_col, c1)

        # extend by borders near content
        border_extra = 0
        for r in range(last_val_row, min(last_val_row + 15, max_scan_row) + 1):
            for c in range(1, last_val_col + 1):
                if _cell_has_border(ws.cell(r, c)):
                    border_extra = max(border_extra, r - last_val_row)
        last_val_row += border_extra

        n_rows = last_val_row
        n_cols = min(max(last_val_col, 1), 100)

        matrix: list[list[str]] = []
        row_border_count = [0] * n_rows
        for r in range(1, n_rows + 1):
            row_vals: list[str] = []
            bcount = 0
            for c in range(1, n_cols + 1):
                cell = ws.cell(r, c)
                text = _format_cell_value(cell.value)
                if not text and wb_formulas is not None:
                    try:
                        fval = wb_formulas[ws.title].cell(r, c).value
                        if isinstance(fval, str) and fval.startswith("="):
                            text = fval
                    except Exception:
                        pass
                row_vals.append(text)
                if _cell_has_border(cell):
                    bcount += 1
            matrix.append(row_vals)
            row_border_count[r - 1] = bcount

        # merges → 0-based relative to matrix (full sheet crop from A1)
        merges: list[tuple[int, int, int, int]] = []
        row_merge_count = [0] * n_rows
        for r0, c0, r1, c1 in merges_abs:
            if r0 > n_rows or c0 > n_cols:
                continue
            rr0 = r0 - 1
            rr1 = min(r1, n_rows) - 1
            cc0 = c0 - 1
            cc1 = min(c1, n_cols) - 1
            if rr1 < 0 or cc1 < 0:
                continue
            merges.append((max(0, rr0), max(0, cc0), rr1, cc1))
            for rr in range(max(0, rr0), rr1 + 1):
                if rr < n_rows:
                    row_merge_count[rr] += 1

        blocks, region_notes = _sheet_to_blocks(
            matrix, merges, row_border_count, row_merge_count
        )
        for note in region_notes:
            warnings.append(f"{ws.title}: {note}")
        text_parts = [b.text for b in blocks if b.text]
        text = "\n".join(text_parts)
        tables = [b for b in blocks if b.kind == "table" and b.grid]
        n_nonempty = sum(
            1
            for b in tables
            for row in (b.grid or [])
            for c in row
            if not c.covered and (c.text or "").strip()
        )
        first_grid = tables[0].grid if tables else []
        sheets.append(
            ExcelSheetModel(
                name=ws.title,
                blocks=blocks,
                text=text,
                n_rows=n_rows,
                n_cols=n_cols,
                n_nonempty=n_nonempty,
                n_tables=len(tables),
                hidden=hidden,
                grid=first_grid or [],
            )
        )
        if text:
            all_text.append(text)

        if n_rows >= 200:
            warnings.append(f"{ws.title}: {n_rows} used rows (max_row={ws.max_row})")

    wb.close()
    if wb_formulas is not None:
        wb_formulas.close()

    return ExcelWorkbookModel(
        sheets=sheets,
        text="\n\n".join(all_text),
        warnings=warnings,
        source_format="xlsx",
    )


def extract_excel(path: str | Path, *, fmt: str | None = None) -> ExcelWorkbookModel:
    """Только .xlsx/.xlsm."""
    src = Path(path)
    kind = (fmt or src.suffix.lstrip(".")).lower()
    if kind == "xls":
        raise RuntimeError(".xls is routed earlier, not extracted here")
    return extract_xlsx(src)
